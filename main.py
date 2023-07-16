from threading import *
import pandas as pd
import logging
import random
import queue
import time
import csv
import re
from collections import defaultdict
import timeit
import openpyxl

import producer
import consumer
import write
from enum import Enum

class LogStatus(Enum):
    read=7
    write=8
    process=9

def logToXLSX(record, i):
    base_row = 2
    wb2 = openpyxl.load_workbook('C:\ofile\\bnchmark.xlsx')
    sheet = wb2.active
    if(record['type']=='read'):
        sheet.cell(row=base_row+i, column = LogStatus.read.value).value=record['value']
    elif(record['type']=='write'):
        sheet.cell(row=base_row+i, column = LogStatus.write.value).value=record['value']
    elif(record['type']=='process'):
        sheet.cell(row=base_row+i, column = LogStatus.process.value).value=record['value']
    wb2.save('C:\ofile\\bnchmark.xlsx')


if __name__ == '__main__':

    for x in range(1,30):
        print(x)
        sharedDic = {}
        p = producer.ProducerThread(name='producer', STOP=object , Nrows=x*10000,sharedDic=sharedDic)
        c = consumer.ConsumerThread(name='consumer', STOP=object ,n=x*1000,sharedDic=sharedDic)
        d = write.ConsumerTh(name="consumer", STOP=object,sharedDic=sharedDic)
        f = write.Consumerunhealthy(name="consumer", STOP=object,sharedDic=sharedDic)
        p.start()
        c.start()
        d.start()
        f.start()
        while (len(sharedDic)<3) :
            time.sleep(1)
        p.join()
        c.join()
        d.join()
        f.join()
        logToXLSX({"type":"read","value":sharedDic["read"]},x)
        logToXLSX({"type": "write", "value": sharedDic["write"]}, x)
        logToXLSX({"type": "process", "value": sharedDic["process"]}, x)
        print("read time",sharedDic["read"])
        print("write time", sharedDic["write"])
        print("process time", sharedDic["process"])







